"""ETL: แปลงข้อมูล SMI-V (HDC export) + ประชากรกลางปี 2567 เป็น docs/smiv-data.json

SMI-V = ผู้ป่วยจิตเวชและสารเสพติดที่มีความเสี่ยงสูงต่อการก่อความรุนแรง (1B030-33)
ข้อมูลเป็นยอดสะสมรายจังหวัด (ไม่มีมิติเวลา/อำเภอ)
"""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from parse import PROVINCES, parse_population

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
OUTPUT_PATH = PROJECT_ROOT / "docs" / "smiv-data.json"

SMIV_FILE = "export_data_20260709_1783610942_H763Y924CA.xlsx"
SMIV_SHEET = "Sheet1"
POP2567_FILE = "ข้อมูลประชากรกลางปี-2567 (1).xlsx"

SMIV_DATA_START_ROW = 3  # แถวข้อมูลจังหวัดแรก (0-indexed)

# คอลัมน์ตามหัวตาราง B-O ของรายงาน HDC
COL_AREA = 0
COL_OLD = 1            # B รายเก่าก่อนปีงบ
COL_NEW = 2            # C รายใหม่ปีงบปัจจุบัน
COL_TOTAL = 3          # D ทั้งหมด (B+C)
COL_ACCESS_RATE = 4    # E อัตราเข้าถึงบริการ (D/I x100)
COL_NO_REPEAT = 5      # F ไม่ก่อความรุนแรงซ้ำ (สะสม)
COL_KPI = 6            # G ร้อยละดูแลต่อเนื่อง+ไม่ก่อซ้ำ (N/I x100)
COL_POP_15_60 = 7      # H ประชากรกลางปี 15-60 ปี (ย้อนหลัง 2 ปี)
COL_ESTIMATED = 8      # I ประมาณการผู้ป่วยในพื้นที่
COL_FOLLOW1 = 9        # J ติดตาม 1 ครั้ง
COL_FOLLOW1_NO_REPEAT = 10  # K ติดตาม 1 ครั้ง + ไม่ก่อซ้ำ
COL_FOLLOW2 = 12       # M ติดตามอย่างน้อย 2 ครั้ง
COL_FOLLOW2_NO_REPEAT = 13  # N ติดตามอย่างน้อย 2 ครั้ง + ไม่ก่อซ้ำ

RATE_TOLERANCE = 0.1  # เผื่อการปัดเศษของร้อยละในไฟล์ต้นฉบับ

# เป้าหมายตามเอกสารชี้แจงตัวชี้วัด SMI-V ปีงบ 2569
TARGETS = {"round6": 35.0, "round9": 37.0, "round12": 40.0}
CURRENT_ROUND = "round9"  # ข้อมูล ณ 9 ก.ค. 2569 ~ รอบ 9 เดือน


def parse_smiv(path: Path) -> dict[str, dict]:
    """อ่านข้อมูล SMI-V รายจังหวัด + แถวรวม พร้อมตรวจสอบความสอดคล้อง"""
    if not path.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ SMI-V: {path}")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SMIV_SHEET not in workbook.sheetnames:
        raise ValueError(f"ไม่พบ sheet '{SMIV_SHEET}' ในไฟล์ SMI-V")
    rows = list(workbook[SMIV_SHEET].iter_rows(values_only=True))

    result: dict[str, dict] = {}
    for row in rows[SMIV_DATA_START_ROW:]:
        area = str(row[COL_AREA]).strip() if row[COL_AREA] else ""
        if not area:
            continue
        key = "รวม" if area == "รวม" else area
        if key != "รวม" and key not in PROVINCES:
            raise ValueError(f"พบพื้นที่ที่ไม่รู้จักในไฟล์ SMI-V: '{area}'")
        entry = {
            "old": int(row[COL_OLD]),
            "new": int(row[COL_NEW]),
            "total": int(row[COL_TOTAL]),
            "accessRate": float(row[COL_ACCESS_RATE]),
            "noRepeat": int(row[COL_NO_REPEAT]),
            "kpi": float(row[COL_KPI]),
            "pop1560": int(row[COL_POP_15_60]),
            "estimated": int(row[COL_ESTIMATED]),
            "follow1": int(row[COL_FOLLOW1]),
            "follow1NoRepeat": int(row[COL_FOLLOW1_NO_REPEAT]),
            "follow2": int(row[COL_FOLLOW2]),
            "follow2NoRepeat": int(row[COL_FOLLOW2_NO_REPEAT]),
        }
        validate_entry(area, entry)
        result[key] = entry

    missing = [p for p in PROVINCES if p not in result]
    if missing:
        raise ValueError(f"ข้อมูล SMI-V ไม่ครบทุกจังหวัด ขาด: {missing}")
    if "รวม" not in result:
        raise ValueError("ไม่พบแถวรวมในไฟล์ SMI-V")
    return result


def validate_entry(area: str, e: dict) -> None:
    """ตรวจความสอดคล้องภายในแถว: D=B+C และร้อยละตรงกับสูตรในหัวตาราง"""
    if e["total"] != e["old"] + e["new"]:
        raise ValueError(f"{area}: ทั้งหมด ({e['total']}) != เก่า+ใหม่")
    expected_access = e["total"] / e["estimated"] * 100
    if abs(expected_access - e["accessRate"]) > RATE_TOLERANCE:
        raise ValueError(
            f"{area}: อัตราเข้าถึง {e['accessRate']} ไม่ตรงสูตร D/I ({expected_access:.2f})")
    expected_kpi = e["follow2NoRepeat"] / e["estimated"] * 100
    if abs(expected_kpi - e["kpi"]) > RATE_TOLERANCE:
        raise ValueError(
            f"{area}: KPI {e['kpi']} ไม่ตรงสูตร N/I ({expected_kpi:.2f})")


def main() -> int:
    smiv = parse_smiv(RAW_DIR / SMIV_FILE)
    population = parse_population(RAW_DIR / POP2567_FILE)

    output = {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "fiscalYear": 2569,
            "asOfLabel": "ข้อมูลสะสมปีงบประมาณ 2569 (ณ 9 กรกฎาคม 2569)",
            "targets": TARGETS,
            "currentRound": CURRENT_ROUND,
            "populationYear": 2567,
            "source": "ระบบคลังข้อมูลด้านการแพทย์และสุขภาพ (HDC) และประชากรกลางปี 2567 กองยุทธศาสตร์และแผนงาน",
        },
        "provinces": PROVINCES,
        "population": population,
        "data": smiv,
    }
    OUTPUT_PATH.write_text(
        json.dumps(output, ensure_ascii=False, indent=1), encoding="utf-8")

    total = smiv["รวม"]
    print(f"ผู้ป่วย SMI-V ทั้งหมด {total['total']:,} คน (ประมาณการ {total['estimated']:,})")
    print(f"KPI ดูแลต่อเนื่อง+ไม่ก่อซ้ำ: {total['kpi']}% (เป้ารอบ 9 เดือน ≥ {TARGETS['round9']}%)")
    print(f"อัตราเข้าถึงบริการ: {total['accessRate']}%")
    print(f"ประชากรกลางปี 2567 เขต 8: {population['รวม']:,}")
    print(f"เขียนผลลัพธ์: {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
