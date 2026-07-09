"""ETL: แปลงรายงานเฝ้าระวังการทำร้ายตนเอง (รง.506S, HTML-in-.xls) และ
ข้อมูลประชากรกลางปี 2568 (xlsx) เป็นข้อมูลสรุปรวม docs/data.json

สำคัญ: output เป็นข้อมูลสรุปรวม (aggregate) เท่านั้น ไม่มีข้อมูลรายบุคคล
"""

from __future__ import annotations

import html
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
OUTPUT_PATH = PROJECT_ROOT / "docs" / "data.json"

CASE_FILES = [
    "1ตค69-31ธค69.xls",
    "1มค69-31มีค69.xls",
    "1เมย69-9กค69.xls",
]
POPULATION_FILE = "ข้อมูลประชากรกลางปี-2568.xlsx"
POPULATION_SHEET = "เขต8"
HDC_FILE = "export_data_20260709_1783566659_WLG3GIXS58.xlsx"
HDC_SHEET = "Sheet1"

# ตำแหน่งคอลัมน์ในรายงาน 506S (202 คอลัมน์)
COL_PROVINCE = 2
COL_AGE = 5
COL_SEX = 6
COL_MONTH = 15
COL_YEAR = 16
COL_OUTCOME = 42
COL_ACCESS_7_2 = 187  # 7.2 การตรวจประเมินตามมาตรฐานจิตเวชและการช่วยเหลือทางสังคมจิตใจ
EXPECTED_COLUMNS = 202

METHOD_COLUMNS = {
    27: "แขวนคอ",
    28: "ของมีคม/ของแข็ง",
    29: "ปืน/ระเบิด",
    30: "กระโดดจากที่สูง",
    31: "จมน้ำ",
    32: "ให้รถชน",
    33: "รมควัน/แก๊ส",
    34: "ขับรถชน",
    35: "สารพิษ",
    37: "กินยาเกินขนาด",
    39: "อื่นๆ",
}

# กลุ่มเสี่ยง: ปัจจัยกระตุ้น (Cri.) และปัจจัยเสี่ยง R1-R10 ของแบบ 506S
RISK_COLUMNS = {
    45: "ประสบปัญหาชีวิต",
    91: "ป่วยด้วยโรคจิตเวช",
    105: "ติดสุรา",
    107: "ติดสารเสพติด",
    109: "โรคทางกายเรื้อรัง",
    131: "บุคลิกภาพเสี่ยง",
    136: "เคยทำร้ายตนเองมาก่อน",
    138: "ครอบครัวเคยฆ่าตัวตาย",
    140: "ประสบการณ์เลวร้ายวัยเด็ก",
    142: "ค่านิยม/ความเชื่อส่วนบุคคล",
    144: "ปัจจัยเสี่ยงอื่นๆ",
}

OUTCOME_DEATH = "ตาย"
OUTCOME_ATTEMPTS = ("บาดเจ็บ", "ไม่บาดเจ็บ")

PROVINCES = ["เลย", "หนองบัวลำภู", "อุดรธานี", "หนองคาย", "บึงกาฬ", "สกลนคร", "นครพนม"]

AGE_GROUPS = ["0-14", "15-24", "25-34", "35-44", "45-54", "55-64", "65+", "ไม่ระบุ"]

# ปีงบประมาณ 2569: ต.ค. 2568 - ก.ย. 2569 (ข้อมูลถึง 9 ก.ค. 2569)
FISCAL_MONTHS = [
    "2568-10", "2568-11", "2568-12",
    "2569-01", "2569-02", "2569-03",
    "2569-04", "2569-05", "2569-06", "2569-07",
]
PERIOD_START = date(2025, 10, 1)
PERIOD_END = date(2026, 7, 9)

KPI1_TARGET = 10.0  # อัตราฆ่าตัวตายสำเร็จ ไม่เกิน 10 ต่อแสนประชากร
KPI2_TARGET = 70.0  # ร้อยละการเข้าถึงบริการของผู้พยายามฆ่าตัวตาย

# โครงสร้างไฟล์ HDC: แถวข้อมูลจังหวัดเริ่มแถว 4, เดือน m เริ่มคอลัมน์ 5 + m*12
# แต่ละเดือนมี 6 กลุ่มอายุ x (คน, ครั้ง) — นับเฉพาะคอลัมน์ "คน"
HDC_DATA_START_ROW = 4
HDC_MONTH_START_COL = 5
HDC_COLS_PER_MONTH = 12
HDC_AGE_BUCKETS = 6
HDC_MONTH_KEYS = [
    "2568-10", "2568-11", "2568-12",
    "2569-01", "2569-02", "2569-03",
    "2569-04", "2569-05", "2569-06", "2569-07",
    None, None,  # ส.ค. / ก.ย. อยู่นอกช่วงข้อมูล (ค่าเป็นศูนย์)
]
HDC_COL_MALE_YEAR = 1
HDC_COL_FEMALE_YEAR = 2

# จำนวนวันของแต่ละเดือนในช่วงข้อมูล (ก.ค. 2569 มีถึงวันที่ 9)
MONTH_DAYS = {
    "2568-10": 31, "2568-11": 30, "2568-12": 31,
    "2569-01": 31, "2569-02": 28, "2569-03": 31,
    "2569-04": 30, "2569-05": 31, "2569-06": 30, "2569-07": 9,
}


def strip_tags(fragment: str) -> str:
    """ลบ HTML tag และ decode entity ให้เหลือข้อความล้วน"""
    return html.unescape(re.sub(r"<[^>]+>", "", fragment)).strip()


def parse_case_file(path: Path) -> list[list[str]]:
    """อ่านไฟล์รายงาน 506S (HTML table) คืนค่าเป็น list ของแถวข้อมูล"""
    if not path.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ข้อมูล: {path}")
    content = path.read_text(encoding="utf-8")
    content = re.sub(r"<!--.*?-->", "", content, flags=re.S)
    records = []
    for row_html in re.findall(r"<tr[^>]*>(.*?)</tr>", content, flags=re.S):
        cells = re.findall(r"<td[^>]*>(.*?)</td>", row_html, flags=re.S)
        if not cells:
            continue  # แถว header (มีแต่ <th>)
        if len(cells) != EXPECTED_COLUMNS:
            raise ValueError(
                f"{path.name}: พบแถวที่มี {len(cells)} คอลัมน์ "
                f"(คาดหวัง {EXPECTED_COLUMNS})"
            )
        records.append([strip_tags(c) for c in cells])
    if not records:
        raise ValueError(f"{path.name}: ไม่พบข้อมูลในไฟล์")
    return records


def age_group(age_text: str) -> str:
    """จัดกลุ่มอายุแบบหยาบเพื่อความเป็นส่วนตัวของข้อมูล"""
    try:
        age = int(age_text)
    except (ValueError, TypeError):
        return "ไม่ระบุ"
    if age < 0:
        return "ไม่ระบุ"
    if age <= 14:
        return "0-14"
    if age <= 24:
        return "15-24"
    if age <= 34:
        return "25-34"
    if age <= 44:
        return "35-44"
    if age <= 54:
        return "45-54"
    if age <= 64:
        return "55-64"
    return "65+"


def month_key(record: list[str]) -> str:
    """สร้าง key ปี-เดือน เช่น '2568-10' จากคอลัมน์ปี/เดือน"""
    year = record[COL_YEAR].strip()
    month = record[COL_MONTH].strip().zfill(2)
    return f"{year}-{month}"


def empty_bucket() -> dict:
    return {
        "die": 0,
        "attempt": 0,
        "access": 0,
        "bySex": {},
        "byAge": {},
        "byMethod": {},
        "byRisk": {},
    }


def aggregate(records: list[list[str]]) -> dict:
    """สรุปข้อมูลรายเคสเป็นยอดรวมราย จังหวัด x เดือน (ไม่มีข้อมูลรายบุคคล)"""
    result: dict[str, dict[str, dict]] = {}
    for rec in records:
        province = rec[COL_PROVINCE].strip()
        if province not in PROVINCES:
            raise ValueError(f"พบจังหวัดนอกเขตสุขภาพที่ 8: '{province}'")
        month = month_key(rec)
        if month not in FISCAL_MONTHS:
            raise ValueError(f"พบเดือนนอกช่วงข้อมูล: '{month}'")

        bucket = result.setdefault(province, {}).setdefault(month, empty_bucket())
        outcome = rec[COL_OUTCOME].strip()
        sex = rec[COL_SEX].strip() or "ไม่ระบุ"
        group = age_group(rec[COL_AGE])

        is_death = outcome == OUTCOME_DEATH
        is_attempt = outcome in OUTCOME_ATTEMPTS
        if not is_death and not is_attempt:
            raise ValueError(f"พบผลการกระทำที่ไม่รู้จัก: '{outcome}'")

        kind = "die" if is_death else "attempt"
        bucket[kind] += 1
        if is_attempt and rec[COL_ACCESS_7_2].strip() == "มี":
            bucket["access"] += 1

        sex_bucket = bucket["bySex"].setdefault(sex, {"die": 0, "attempt": 0})
        sex_bucket[kind] += 1
        age_bucket = bucket["byAge"].setdefault(group, {"die": 0, "attempt": 0})
        age_bucket[kind] += 1
        for col, method_name in METHOD_COLUMNS.items():
            if rec[col].strip() == "มี":
                bucket["byMethod"][method_name] = (
                    bucket["byMethod"].get(method_name, 0) + 1
                )
        for col, risk_name in RISK_COLUMNS.items():
            if rec[col].strip() == "มี":
                bucket["byRisk"][risk_name] = (
                    bucket["byRisk"].get(risk_name, 0) + 1
                )
    return result


def parse_population(path: Path) -> dict[str, int]:
    """อ่านประชากรกลางปี 2568 จาก sheet เขต8 (โครงสร้าง 3 บล็อก บล็อกละ 3 พื้นที่)"""
    if not path.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ประชากร: {path}")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if POPULATION_SHEET not in workbook.sheetnames:
        raise ValueError(f"ไม่พบ sheet '{POPULATION_SHEET}' ในไฟล์ประชากร")
    rows = list(workbook[POPULATION_SHEET].iter_rows(values_only=True))

    # แต่ละบล็อก: (แถวชื่อจังหวัด, แถว TOTAL) / แต่ละพื้นที่: (คอลัมน์ชื่อ, คอลัมน์รวม)
    blocks = [(1, 25), (29, 53), (57, 81)]
    area_columns = [(2, 3), (5, 6), (8, 9)]

    population: dict[str, int] = {}
    for header_row, total_row in blocks:
        for name_col, total_col in area_columns:
            name = rows[header_row][name_col]
            if name is None:
                continue
            name = str(name).strip()
            total = rows[total_row][total_col]
            if total is None:
                raise ValueError(f"ไม่พบยอดประชากรของ '{name}'")
            key = "รวม" if "รวมเขต" in name else name
            population[key] = int(total)

    missing = [p for p in PROVINCES if p not in population]
    if missing:
        raise ValueError(f"ประชากรไม่ครบทุกจังหวัด ขาด: {missing}")
    province_sum = sum(population[p] for p in PROVINCES)
    if province_sum != population["รวม"]:
        raise ValueError(
            f"ยอดรวมประชากรไม่ตรง: รายจังหวัดรวม {province_sum:,} "
            f"แต่ยอดเขต {population['รวม']:,}"
        )
    return population


def parse_hdc(path: Path) -> dict:
    """อ่านจำนวนผู้พยายามฆ่าตัวตายที่เข้าถึงบริการจาก HDC (คน รายจังหวัด x เดือน)

    หมายเหตุ: ยอด "yearly" คือจำนวนคนไม่ซ้ำทั้งปี ส่วนยอดรายเดือนนับคนไม่ซ้ำ
    ภายในเดือนนั้น (ผลรวมรายเดือนจึงอาจมากกว่ายอดทั้งปี)
    """
    if not path.exists():
        raise FileNotFoundError(f"ไม่พบไฟล์ HDC: {path}")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if HDC_SHEET not in workbook.sheetnames:
        raise ValueError(f"ไม่พบ sheet '{HDC_SHEET}' ในไฟล์ HDC")
    rows = list(workbook[HDC_SHEET].iter_rows(values_only=True))

    by_prov_month: dict[str, dict[str, int]] = {}
    yearly: dict[str, int] = {}
    for row in rows[HDC_DATA_START_ROW:]:
        province = str(row[0]).strip() if row[0] else ""
        if province not in PROVINCES:
            continue  # ข้ามแถวรวมท้ายตารางและแถวว่าง
        yearly[province] = int(row[HDC_COL_MALE_YEAR] or 0) + int(row[HDC_COL_FEMALE_YEAR] or 0)
        monthly: dict[str, int] = {}
        for m, month_key in enumerate(HDC_MONTH_KEYS):
            if month_key is None:
                continue
            start = HDC_MONTH_START_COL + m * HDC_COLS_PER_MONTH
            monthly[month_key] = sum(
                int(row[start + i * 2] or 0) for i in range(HDC_AGE_BUCKETS)
            )
        by_prov_month[province] = monthly

    missing = [p for p in PROVINCES if p not in yearly]
    if missing:
        raise ValueError(f"ข้อมูล HDC ไม่ครบทุกจังหวัด ขาด: {missing}")
    yearly["รวม"] = sum(yearly[p] for p in PROVINCES)
    return {"byProvMonth": by_prov_month, "yearly": yearly}


def validate(data: dict, expected_total: int) -> None:
    """ตรวจสอบยอดรวมหลัง aggregate ต้องเท่ากับจำนวนเคสต้นฉบับ"""
    total = sum(
        bucket["die"] + bucket["attempt"]
        for months in data.values()
        for bucket in months.values()
    )
    if total != expected_total:
        raise ValueError(f"ยอดรวมไม่ตรง: aggregate ได้ {total} ต้นฉบับ {expected_total}")


def build_output(data: dict, population: dict[str, int], hdc: dict) -> dict:
    days_covered = (PERIOD_END - PERIOD_START).days + 1
    return {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "fiscalYear": 2569,
            "periodLabel": "1 ตุลาคม 2568 – 9 กรกฎาคม 2569",
            "daysCovered": days_covered,
            "kpi1Target": KPI1_TARGET,
            "kpi2Target": KPI2_TARGET,
            "populationYear": 2568,
            "source": "รายงานเฝ้าระวังการทำร้ายตนเอง (รง.506S), ข้อมูลการเข้าถึงบริการจาก HDC และประชากรกลางปี 2568 กองยุทธศาสตร์และแผนงาน",
        },
        "months": FISCAL_MONTHS,
        "monthDays": MONTH_DAYS,
        "provinces": PROVINCES,
        "ageGroups": AGE_GROUPS,
        "population": population,
        "hdc": hdc,
        "data": data,
    }


def main() -> int:
    all_records: list[list[str]] = []
    for filename in CASE_FILES:
        records = parse_case_file(RAW_DIR / filename)
        print(f"อ่าน {filename}: {len(records)} เคส")
        all_records.extend(records)

    data = aggregate(all_records)
    validate(data, len(all_records))
    population = parse_population(RAW_DIR / POPULATION_FILE)
    hdc = parse_hdc(RAW_DIR / HDC_FILE)

    output = build_output(data, population, hdc)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(output, ensure_ascii=False, indent=1), encoding="utf-8"
    )

    deaths = sum(b["die"] for m in data.values() for b in m.values())
    attempts = sum(b["attempt"] for m in data.values() for b in m.values())
    assessed = sum(b["access"] for m in data.values() for b in m.values())
    hdc_total = hdc["yearly"]["รวม"]
    rate = deaths / population["รวม"] * 100_000
    print(f"\nรวม {len(all_records)} เคส | เสียชีวิต {deaths} | พยายาม {attempts}")
    print(f"KPI1 อัตราฆ่าตัวตายสำเร็จสะสม: {rate:.2f} ต่อแสน (เป้า ≤ {KPI1_TARGET})")
    kpi2 = min(100.0, attempts / hdc_total * 100)
    print(f"KPI2 เข้าถึงบริการ = 100 x (506S/HDC): {attempts}/{hdc_total} = {kpi2:.1f}% (เป้า ≥ {KPI2_TARGET}%)")
    print(f"ตัวชี้วัดรอง ตรวจประเมินจิตเวช (ข้อ 7.2): {assessed}/{attempts} = {assessed / attempts * 100:.1f}%")
    print(f"\nเขียนผลลัพธ์: {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
